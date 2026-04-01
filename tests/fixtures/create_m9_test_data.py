from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

ROW_COUNT = 50
SPECIES_COLUMNS = ["A", "B", "C", "Int", "Water", "PBMPBPIN", "PBMP", "pinBOH"]


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _load_source_rows(filename: str) -> list[dict[str, str]]:
    path = _repo_root() / "tests" / "data" / "synthetic" / "complex_mechanism_global" / filename
    with open(path, "r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
    if len(rows) < ROW_COUNT:
        raise ValueError(f"Expected at least {ROW_COUNT} rows in {path}, found {len(rows)}.")
    return rows[:ROW_COUNT]


def _scaled_time_series(max_value: float) -> list[float]:
    return [round(index * max_value / (ROW_COUNT - 1), 9) for index in range(ROW_COUNT)]


def _scaled_species_rows(source_rows: list[dict[str, str]], concentration_scale: float) -> list[list[float]]:
    scaled_rows: list[list[float]] = []
    for row in source_rows:
        scaled_rows.append([
            round(float(row[column]) * concentration_scale, 9)
            for column in SPECIES_COLUMNS
        ])
    return scaled_rows


def _sheet_rows(
    *,
    time_values: list[float],
    species_rows: list[list[float]],
    notes: list[str] | None = None,
    temperatures: list[str] | None = None,
) -> list[list[object]]:
    rows: list[list[object]] = []
    for index, (time_value, species_values) in enumerate(zip(time_values, species_rows)):
        row: list[object] = [time_value, *species_values]
        if notes is not None:
            row.append(notes[index])
        if temperatures is not None:
            row.append(temperatures[index])
        rows.append(row)
    return rows


def _write_sheet(workbook: Workbook, title: str, header: list[str], unit_row: list[str] | None, rows: list[list[object]]) -> None:
    worksheet = workbook.active if workbook.active.title == "Sheet" and workbook.active.max_row == 1 else workbook.create_sheet()
    worksheet.title = title
    worksheet.append(header)
    if unit_row is not None:
        worksheet.append(unit_row)
    for row in rows:
        worksheet.append(row)


def create_workbook(output_path: Path) -> Path:
    pH7_source = _load_source_rows("dataset_01.csv")
    pH9_source = _load_source_rows("dataset_06.csv")

    pH7_time_us = _scaled_time_series(500.0)
    pH9_time_us = _scaled_time_series(500.0)
    pH7_species_uM = _scaled_species_rows(pH7_source, concentration_scale=1_000.0)
    pH9_species_uM = _scaled_species_rows(pH9_source, concentration_scale=1_000.0)

    workbook = Workbook()

    full_header = ["time", *SPECIES_COLUMNS]
    full_unit_row_um = ["us", *(["uM"] * len(SPECIES_COLUMNS))]
    _write_sheet(
        workbook,
        "pH7_run",
        full_header,
        full_unit_row_um,
        _sheet_rows(time_values=pH7_time_us, species_rows=pH7_species_uM),
    )
    _write_sheet(
        workbook,
        "pH9_run",
        full_header,
        full_unit_row_um,
        _sheet_rows(time_values=pH9_time_us, species_rows=pH9_species_uM),
    )

    _write_sheet(
        workbook,
        "different_units",
        full_header,
        ["ms", *(["nM"] * len(SPECIES_COLUMNS))],
        _sheet_rows(
            time_values=[round(value / 1_000.0, 9) for value in pH7_time_us],
            species_rows=_scaled_species_rows(pH7_source, concentration_scale=1_000_000.0),
        ),
    )

    _write_sheet(
        workbook,
        "no_unit_row",
        full_header,
        None,
        _sheet_rows(
            time_values=[round(value * 1e-6, 9) for value in pH7_time_us],
            species_rows=_scaled_species_rows(pH7_source, concentration_scale=1e-3),
        ),
    )

    notes = [
        ["baseline", "mix", "steady", "late"][index % 4]
        for index in range(ROW_COUNT)
    ]
    temperatures = [
        ["298K", "299K", "300K"][index % 3]
        for index in range(ROW_COUNT)
    ]
    _write_sheet(
        workbook,
        "extra_columns",
        [*full_header, "notes", "temperature"],
        [*full_unit_row_um, "", ""],
        _sheet_rows(
            time_values=pH7_time_us,
            species_rows=pH7_species_uM,
            notes=notes,
            temperatures=temperatures,
        ),
    )

    different_structure_rows: list[list[object]] = []
    for time_value, species_values in zip(pH7_time_us, pH7_species_uM):
        x_value = round(species_values[3] + species_values[5], 9)
        y_value = round(species_values[6] + species_values[7], 9)
        different_structure_rows.append([time_value, x_value, y_value])
    _write_sheet(
        workbook,
        "different_structure",
        ["time", "X", "Y"],
        ["us", "uM", "uM"],
        different_structure_rows,
    )

    parametric_rows: list[list[object]] = []
    for time_value, species_values in zip(pH7_time_us, pH7_species_uM):
        observable_x = round(species_values[5] + species_values[6], 9)
        parametric_rows.append([
            round(time_value * 1e-6, 9),
            observable_x,
            species_values[0],
            species_values[1],
        ])
    _write_sheet(
        workbook,
        "parametric_x",
        ["time", "observable_X", "A", "B"],
        ["s", "uM", "uM", "uM"],
        parametric_rows,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return output_path


def main() -> None:
    output_path = _repo_root() / "tests" / "fixtures" / "m9_test_datasets.xlsx"
    created = create_workbook(output_path)
    print(created)


if __name__ == "__main__":
    main()
