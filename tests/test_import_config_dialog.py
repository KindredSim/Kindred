"""Tests for the import configuration dialog.

Covers auto-detection (time column, unit row), Excel sheet handling,
ImportConfig return values, dialog actions, validation, unit-row override,
preview content, and the apply-to-remaining checkbox.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List

import pytest
from PySide6 import QtCore, QtWidgets

from kindred.gui.widgets.import_config_dialog import ImportConfigDialog

pytestmark = pytest.mark.gui


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_csv(path, header: List[str], rows: List[List[str]]) -> str:
    """Write a CSV file and return its path as a string."""
    filepath = str(path)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)
    return filepath


def _write_xlsx(path, sheets: dict) -> str:
    """Write an Excel workbook with {sheet_name: (header, rows)} and return path."""
    from openpyxl import Workbook

    filepath = str(path)
    wb = Workbook()
    first = True
    for sheet_name, (header, rows) in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        ws.append(header)
        for row in rows:
            ws.append(row)
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# 1. CSV auto-detection
# ---------------------------------------------------------------------------

class TestCsvAutoDetection:
    """CSV time-column and unit-row auto-detection."""

    def test_time_column_detected(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "a.csv", ["time", "A", "B"], [
            ["0", "1.0", "2.0"],
            ["1", "1.1", "2.1"],
        ])
        dlg = ImportConfigDialog(fp)
        assert dlg._time_combo.currentText() == "time"

    def test_t_column_detected(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "b.csv", ["t", "X", "Y"], [
            ["0", "1.0", "2.0"],
        ])
        dlg = ImportConfigDialog(fp)
        assert dlg._time_combo.currentText() == "t"

    def test_no_standard_time_column(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "c.csv", ["elapsed", "A", "B"], [
            ["0", "1.0", "2.0"],
        ])
        dlg = ImportConfigDialog(fp)
        # No auto-detection: combo should show first column but no pre-selection
        # The import button should be disabled because no valid time column is confirmed
        # Actually per spec: "no column is pre-selected — user must pick manually"
        assert dlg._time_combo.currentText() == ""

    def test_unit_row_detected(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "d.csv", ["time", "A", "B"], [
            ["s", "uM", "uM"],
            ["0", "1.0", "2.0"],
            ["1", "1.1", "2.1"],
        ])
        dlg = ImportConfigDialog(fp)
        assert dlg._unit_row_detected is True
        assert dlg._time_unit_combo.currentText() == "s"
        assert dlg._conc_unit_combo.currentText() == "uM"

    def test_unicode_micro_units_normalize_to_ascii_combo_entries(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "unicode_units.csv", ["time", "A", "B"], [
            ["µs", "µM", "μM"],
            ["0", "1.0", "2.0"],
            ["1", "1.1", "2.1"],
        ])
        dlg = ImportConfigDialog(fp)
        assert dlg._time_unit_combo.currentText() == "us"
        assert dlg._conc_unit_combo.currentText() == "uM"

    def test_no_unit_row(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "e.csv", ["time", "A", "B"], [
            ["0", "1.0", "2.0"],
            ["1", "1.1", "2.1"],
        ])
        dlg = ImportConfigDialog(fp)
        assert dlg._unit_row_detected is False
        # Defaults
        assert dlg._time_unit_combo.currentText() == "s"
        assert dlg._conc_unit_combo.currentText() == "M"


# ---------------------------------------------------------------------------
# 2. Excel sheet handling
# ---------------------------------------------------------------------------

class TestExcelSheetHandling:
    """Excel multi-sheet listing, preview switching, and validation."""

    def test_multi_sheet_listed_all_checked(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "wb.xlsx", {
            "pH7": (["time", "A"], [["0", "1.0"], ["1", "0.9"]]),
            "pH9": (["time", "A"], [["0", "1.5"], ["1", "1.4"]]),
            "meta": (["key", "val"], [["temp", "298"]]),
        })
        dlg = ImportConfigDialog(fp)
        names = dlg._get_checked_sheet_names()
        assert set(names) == {"pH7", "pH9", "meta"}

    def test_sheet_selection_changes_preview(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "wb2.xlsx", {
            "Run1": (["time", "X"], [["0", "10"], ["1", "20"]]),
            "Run2": (["time", "Y"], [["0", "30"], ["1", "40"]]),
        })
        dlg = ImportConfigDialog(fp)
        # Click Run2 to show its preview
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Run2":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break
        # Preview table should now show Run2 columns
        headers = [
            dlg._preview_table.horizontalHeaderItem(c).text()
            for c in range(dlg._preview_table.columnCount())
        ]
        assert "Y" in headers

    def test_sheet_switch_preserves_per_sheet_configuration(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "stateful.xlsx", {
            "Sheet1": (
                ["time", "alt_time", "A", "B"],
                [["ms", "s", "uM", "nM"], ["0", "0", "1.0", "2.0"]],
            ),
            "Sheet2": (
                ["time", "alt_time", "A", "B"],
                [["us", "ms", "mM", "uM"], ["0", "0", "3.0", "4.0"]],
            ),
        })
        dlg = ImportConfigDialog(fp)

        dlg._time_combo.setCurrentText("alt_time")
        dlg._time_unit_combo.setCurrentText("s")
        dlg._conc_unit_combo.setCurrentText("nM")
        for cb in dlg._species_checkboxes:
            if cb.property("column_name") == "B":
                cb.setChecked(False)

        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Sheet2":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Sheet1":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        species = {
            cb.property("column_name"): cb.isChecked()
            for cb in dlg._species_checkboxes
        }
        assert dlg._time_combo.currentText() == "alt_time"
        assert dlg._time_unit_combo.currentText() == "s"
        assert dlg._conc_unit_combo.currentText() == "nM"
        assert species["B"] is False

    def test_apply_checkbox_copies_current_sheet_configuration_at_import(self, qapp, tmp_path, monkeypatch):
        fp = _write_xlsx(tmp_path / "apply_sheets.xlsx", {
            "Sheet1": (
                ["time", "A", "B"],
                [["ms", "uM", "uM"], ["0", "1.0", "2.0"]],
            ),
            "Sheet2": (
                ["time", "A", "B"],
                [["s", "uM", "uM"], ["0", "3.0", "4.0"]],
            ),
        })
        dlg = ImportConfigDialog(fp)

        dlg._time_unit_combo.setCurrentText("ms")
        dlg._conc_unit_combo.setCurrentText("uM")
        for cb in dlg._species_checkboxes:
            if cb.property("column_name") == "B":
                cb.setChecked(False)

        dlg._apply_remaining_cb.setChecked(True)

        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        result = dlg._build_result("import")
        assert not criticals
        assert result.config is not None
        intents = dict(result.config.per_sheet_intents)
        assert intents["Sheet2"].time_unit == "ms"
        assert intents["Sheet2"].concentration_unit == "uM"
        assert "B" not in intents["Sheet2"].species_columns

    def test_apply_checkbox_rejects_mismatched_columns_at_import(self, qapp, tmp_path, monkeypatch):
        fp = _write_xlsx(tmp_path / "apply_sheets_mismatch.xlsx", {
            "Sheet1": (
                ["time", "A"],
                [["ms", "uM"], ["0", "1.0"]],
            ),
            "Sheet2": (
                ["elapsed", "B"],
                [["s", "nM"], ["0", "2.0"]],
            ),
        })
        dlg = ImportConfigDialog(fp)
        dlg._apply_remaining_cb.setChecked(True)

        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        result = dlg._build_result("import")
        assert criticals, "Mismatched columns must produce an error at import"
        assert result.config is None

    def test_apply_checkbox_keeps_no_unit_row_off_when_target_has_no_unit_row(self, qapp, tmp_path, monkeypatch):
        fp = _write_xlsx(tmp_path / "apply_sheets_no_unit_row.xlsx", {
            "Sheet1": (
                ["time", "A"],
                [["ms", "uM"], ["0", "1.0"]],
            ),
            "Sheet2": (
                ["time", "A"],
                [["0", "2.0"], ["1", "3.0"]],
            ),
        })
        dlg = ImportConfigDialog(fp)
        dlg._no_unit_row_cb.setChecked(True)
        dlg._apply_remaining_cb.setChecked(True)

        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        result = dlg._build_result("import")
        assert not criticals
        assert result.config is not None
        intents = dict(result.config.per_sheet_intents)
        assert intents["Sheet2"].override_no_unit_row is False

    def test_sheet_switch_preview_highlighting_uses_target_sheet_detection(self, qapp, tmp_path):
        """Bug 2 regression: preview row highlighting must use the target
        sheet's detection, not the previous sheet's stale state."""
        from kindred.gui.widgets.import_config_dialog import _UNIT_ROW_BG

        fp = _write_xlsx(tmp_path / "mixed_detection.xlsx", {
            "has_units": (["time", "A"], [["ms", "uM"], ["0", "1.0"]]),
            "no_units": (["time", "A"], [["0", "2.0"], ["1", "3.0"]]),
        })
        dlg = ImportConfigDialog(fp)

        assert dlg._unit_row_detected is True

        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "no_units":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        item = dlg._preview_table.item(0, 0)
        assert item is not None
        bg_color = item.background().color()
        assert bg_color != _UNIT_ROW_BG, (
            "Row 0 should not be highlighted as unit row for a sheet without units"
        )

    def test_sheet_switch_shows_each_sheets_detected_units(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "sheet_units.xlsx", {
            "Sheet1": (
                ["time", "A"],
                [["ms", "uM"], ["0", "1.0"]],
            ),
            "Sheet2": (
                ["time", "A"],
                [["us", "mM"], ["0", "2.0"]],
            ),
        })
        dlg = ImportConfigDialog(fp)

        assert dlg._time_unit_combo.currentText() == "ms"
        assert dlg._conc_unit_combo.currentText() == "uM"

        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Sheet2":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        assert dlg._time_unit_combo.currentText() == "us"
        assert dlg._conc_unit_combo.currentText() == "mM"

    def test_no_sheets_checked_disables_import(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "wb3.xlsx", {
            "Sheet1": (["time", "A"], [["0", "1"]]),
        })
        dlg = ImportConfigDialog(fp)
        # Uncheck the only sheet
        item = dlg._sheet_list.item(0)
        item.setCheckState(QtCore.Qt.CheckState.Unchecked)
        dlg._update_import_enabled()
        assert not dlg._btn_import.isEnabled()

    def test_incompatible_checked_sheets_import_with_independent_defaults(self, qapp, tmp_path, monkeypatch):
        fp = _write_xlsx(tmp_path / "mismatch.xlsx", {
            "Run1": (["time", "A"], [["0", "1.0"]]),
            "Run2": (["time", "B", "C"], [["0", "2.0", "3.0"]]),
        })
        dlg = ImportConfigDialog(fp)
        criticals: list[tuple] = []

        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        dlg._on_import()

        assert not criticals
        assert dlg._result is not None
        assert [plan.sheet_name for plan in dlg._result.config.plans] == ["Run1", "Run2"]

    def test_checked_sheet_compatibility_uses_checked_set_not_previewed_sheet(self, qapp, tmp_path, monkeypatch):
        fp = _write_xlsx(tmp_path / "compatible_checked.xlsx", {
            "Run1": (["time", "A"], [["0", "1.0"]]),
            "Run2": (["time", "B"], [["0", "2.0"]]),
            "Run3": (["time", "B"], [["0", "3.0"]]),
        })
        dlg = ImportConfigDialog(fp)

        # Preview Run2 so species reflect checked sheets' columns
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Run2":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        # Uncheck Run1 (which doesn't have column B)
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Run1":
                item.setCheckState(QtCore.Qt.CheckState.Unchecked)

        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        dlg._on_import()

        assert dlg._result is not None
        assert not criticals

    def test_reordered_checked_sheet_headers_are_compatible(self, qapp, tmp_path, monkeypatch):
        fp = _write_xlsx(tmp_path / "reordered.xlsx", {
            "Run1": (["time", "A", "B"], [["0", "1.0", "2.0"]]),
            "Run2": (["B", "time", "A"], [["3.0", "0", "4.0"]]),
        })
        dlg = ImportConfigDialog(fp)
        criticals: list[tuple] = []

        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        dlg._on_import()

        assert not criticals
        assert dlg._result is not None

    def test_excel_preview_stops_after_preview_limit(self, qapp, tmp_path, monkeypatch):
        from kindred.core.datasets import excel_import
        from kindred.gui.widgets import import_config_dialog as dialog_module

        class _PreviewSheet:
            def __init__(self) -> None:
                self.data_rows_yielded = 0

            def iter_rows(self, values_only: bool = True):
                yield ("time", "A")
                for index in range(100):
                    self.data_rows_yielded += 1
                    if self.data_rows_yielded > dialog_module._MAX_PREVIEW_ROWS + 1:
                        raise AssertionError("preview consumed more rows than needed")
                    yield (index, index + 1.0)

        class _FakeWorkbook:
            def __init__(self) -> None:
                self.sheetnames = ["Data"]
                self.closed = 0
                self._sheet = _PreviewSheet()

            def __getitem__(self, name: str):
                assert name == "Data"
                return self._sheet

            def close(self) -> None:
                self.closed += 1

        workbooks: list[_FakeWorkbook] = []

        def _open_workbook(_path: str):
            workbook = _FakeWorkbook()
            workbooks.append(workbook)
            return workbook

        monkeypatch.setattr(excel_import, "_open_workbook", _open_workbook)

        dlg = ImportConfigDialog(str(tmp_path / "preview.xlsx"))

        preview_workbook = workbooks[-1]
        assert dlg._preview_table.rowCount() == dialog_module._MAX_PREVIEW_ROWS
        assert preview_workbook._sheet.data_rows_yielded == dialog_module._MAX_PREVIEW_ROWS + 1
        assert preview_workbook.closed == 1


# ---------------------------------------------------------------------------
# 3. ImportConfig returned values
# ---------------------------------------------------------------------------

class TestImportConfigValues:
    """Verify returned ImportConfig fields are correct."""

    def test_csv_config_fields(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "exp.csv", ["time", "Conc_A", "Conc_B"], [
            ["us", "uM", "uM"],
            ["0", "1.0", "2.0"],
            ["1", "0.9", "1.9"],
        ])
        dlg = ImportConfigDialog(fp)
        result = dlg._build_result("import")
        cfg = result.config
        assert cfg is not None
        assert cfg.filepath == fp
        assert cfg.file_type == "csv"
        assert cfg.file_intent.sheet_names == ()
        assert cfg.file_intent.apply_to_remaining is False
        assert dict(cfg.per_sheet_intents)[None].time_column == "time"
        assert dict(cfg.per_sheet_intents)[None].species_columns == ("Conc_A", "Conc_B")
        assert dict(cfg.per_sheet_intents)[None].time_unit == "us"
        assert dict(cfg.per_sheet_intents)[None].concentration_unit == "uM"
        assert cfg.plans[0].skip_unit_row is True

    def test_excel_config_sheets_populated(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "k.xlsx", {
            "Run1": (["time", "A"], [["0", "1"]]),
            "Run2": (["time", "A"], [["0", "2"]]),
        })
        dlg = ImportConfigDialog(fp)
        result = dlg._build_result("import")
        cfg = result.config
        assert cfg.file_type == "excel"
        assert set(cfg.file_intent.sheet_names) == {"Run1", "Run2"}

    def test_unit_row_detected_flag(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "no_units.csv", ["time", "A"], [
            ["0", "1.0"],
            ["1", "0.9"],
        ])
        dlg = ImportConfigDialog(fp)
        result = dlg._build_result("import")
        assert result.config.plans[0].skip_unit_row is False

    def test_apply_to_remaining_reflects_checkbox(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "r.csv", ["time", "A"], [
            ["0", "1.0"],
        ])
        dlg = ImportConfigDialog(fp, remaining_count=3)
        dlg._apply_remaining_cb.setChecked(True)
        result = dlg._build_result("import")
        assert result.config.file_intent.apply_to_remaining is True


# ---------------------------------------------------------------------------
# 4. Dialog actions
# ---------------------------------------------------------------------------

class TestDialogActions:
    """Import, Skip, Cancel return correct action strings."""

    def test_import_action(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "f.csv", ["time", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp)
        result = dlg._build_result("import")
        assert result.action == "import"
        assert result.config is not None

    def test_skip_action(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "g.csv", ["time", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp)
        result = dlg._build_result("skip")
        assert result.action == "skip"
        assert result.config is None

    def test_cancel_action(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "h.csv", ["time", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp)
        result = dlg._build_result("cancel")
        assert result.action == "cancel"
        assert result.config is None


# ---------------------------------------------------------------------------
# 5. Validation
# ---------------------------------------------------------------------------

class TestValidation:
    """Import button disabled when required selections are missing."""

    def test_no_time_column_disables_import(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "v1.csv", ["elapsed", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp)
        # No time column auto-detected, none selected
        assert not dlg._btn_import.isEnabled()

    def test_no_species_columns_disables_import(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "v2.csv", ["time", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp)
        # Uncheck all species
        for cb in dlg._species_checkboxes:
            cb.setChecked(False)
        dlg._update_import_enabled()
        assert not dlg._btn_import.isEnabled()

    def test_excel_no_sheets_disables_import(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "v3.xlsx", {
            "S1": (["time", "A"], [["0", "1"]]),
        })
        dlg = ImportConfigDialog(fp)
        dlg._sheet_list.item(0).setCheckState(QtCore.Qt.CheckState.Unchecked)
        dlg._update_import_enabled()
        assert not dlg._btn_import.isEnabled()


# ---------------------------------------------------------------------------
# 6. Unit row override
# ---------------------------------------------------------------------------

class TestUnitRowOverride:
    """Checking 'No unit row' clears detection and defaults to s/M."""

    def test_override_clears_units(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "u.csv", ["time", "A"], [
            ["us", "uM"],
            ["0", "1.0"],
        ])
        dlg = ImportConfigDialog(fp)
        assert dlg._unit_row_detected is True
        # Override
        dlg._no_unit_row_cb.setChecked(True)
        result = dlg._build_result("import")
        intent = dict(result.config.per_sheet_intents)[None]
        assert intent.time_unit == "s"
        assert intent.concentration_unit == "M"
        assert result.config.plans[0].skip_unit_row is True
        assert intent.override_no_unit_row is True

    def test_mixed_concentration_units_show_warning_and_allow_override(self, qapp, tmp_path, monkeypatch):
        fp = _write_csv(tmp_path / "mixed_units.csv", ["time", "A", "B"], [
            ["s", "uM", "nM"],
            ["0", "1.0", "2.0"],
        ])
        dlg = ImportConfigDialog(fp)
        assert "Multiple concentration units detected" in dlg._unit_warning_label.text()
        assert dlg._conc_unit_combo.currentText() == "uM"

        # Mixed units are rejected by the resolver
        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )
        result = dlg._build_result("import")
        assert result.config is None

        # Override unit row allows import with default units
        dlg._no_unit_row_cb.setChecked(True)
        result2 = dlg._build_result("import")
        assert result2.config is not None
        assert dict(result2.config.per_sheet_intents)[None].override_no_unit_row is True


# ---------------------------------------------------------------------------
# 7. Preview content
# ---------------------------------------------------------------------------

class TestPreviewContent:
    """Preview table shows correct rows, columns, and unit-row highlighting."""

    def test_preview_shows_rows(self, qapp, tmp_path):
        rows = [["time", "A"]] + [[str(i), str(float(i))] for i in range(5)]
        fp = _write_csv(tmp_path / "p.csv", rows[0], rows[1:])
        dlg = ImportConfigDialog(fp)
        assert dlg._preview_table.rowCount() == 5

    def test_preview_shows_columns(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "q.csv", ["time", "X", "Y"], [["0", "1", "2"]])
        dlg = ImportConfigDialog(fp)
        headers = [
            dlg._preview_table.horizontalHeaderItem(c).text()
            for c in range(dlg._preview_table.columnCount())
        ]
        assert headers == ["time", "X", "Y"]

    def test_unit_row_highlighted(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "hl.csv", ["time", "A"], [
            ["s", "uM"],
            ["0", "1.0"],
        ])
        dlg = ImportConfigDialog(fp)
        # Row 0 in the table is the unit row — should have a distinct background
        item = dlg._preview_table.item(0, 0)
        assert item is not None
        bg = item.background().color()
        # The unit-row background should differ from the default (white/invalid)
        default_bg = QtWidgets.QTableWidgetItem().background().color()
        assert bg != default_bg

    def test_preview_caps_at_20_rows(self, qapp, tmp_path):
        data_rows = [[str(i), str(float(i))] for i in range(30)]
        fp = _write_csv(tmp_path / "big.csv", ["time", "A"], data_rows)
        dlg = ImportConfigDialog(fp)
        # Should show at most ~20 data rows (unit row detection may eat one)
        assert dlg._preview_table.rowCount() <= 20


# ---------------------------------------------------------------------------
# 8. Apply to remaining
# ---------------------------------------------------------------------------

class TestApplyToRemaining:
    """Checkbox visibility and label reflect remaining_count."""

    def test_hidden_when_zero(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "z.csv", ["time", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp, remaining_count=0)
        assert dlg._apply_remaining_cb.isHidden()

    def test_visible_with_remaining(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "z2.csv", ["time", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp, remaining_count=4)
        assert not dlg._apply_remaining_cb.isHidden()
        assert "remaining" in dlg._apply_remaining_cb.text()

    def test_excel_visible_without_remaining(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "wb.xlsx", {
            "S1": (["time", "A"], [["0", "1"]]),
            "S2": (["time", "A"], [["0", "2"]]),
        })
        dlg = ImportConfigDialog(fp, remaining_count=0)
        assert not dlg._apply_remaining_cb.isHidden()
        assert "all other sheets" in dlg._apply_remaining_cb.text()

    def test_excel_with_remaining_label(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "wb.xlsx", {
            "S1": (["time", "A"], [["0", "1"]]),
            "S2": (["time", "A"], [["0", "2"]]),
        })
        dlg = ImportConfigDialog(fp, remaining_count=3)
        assert not dlg._apply_remaining_cb.isHidden()
        assert "all other sheets" in dlg._apply_remaining_cb.text()
        assert "remaining" in dlg._apply_remaining_cb.text()

    def test_checkbox_state_in_config(self, qapp, tmp_path):
        fp = _write_csv(tmp_path / "z3.csv", ["time", "A"], [["0", "1"]])
        dlg = ImportConfigDialog(fp, remaining_count=2)
        dlg._apply_remaining_cb.setChecked(False)
        r1 = dlg._build_result("import")
        assert r1.config.file_intent.apply_to_remaining is False
        dlg._apply_remaining_cb.setChecked(True)
        r2 = dlg._build_result("import")
        assert r2.config.file_intent.apply_to_remaining is True

    def test_unified_apply_checkbox_copies_sheet_state_at_import(self, qapp, tmp_path, monkeypatch):
        """UX merge regression: checking the apply checkbox and importing must
        copy the current sheet's configuration to all other checked sheets."""
        fp = _write_xlsx(tmp_path / "multi.xlsx", {
            "Sheet1": (["time", "A", "B"], [["ms", "uM", "uM"], ["0", "1.0", "2.0"]]),
            "Sheet2": (["time", "A", "B"], [["s", "uM", "uM"], ["0", "3.0", "4.0"]]),
        })
        dlg = ImportConfigDialog(fp, remaining_count=2)

        for cb in dlg._species_checkboxes:
            if cb.property("column_name") == "B":
                cb.setChecked(False)

        dlg._apply_remaining_cb.setChecked(True)

        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        result = dlg._build_result("import")
        assert not criticals
        assert result.config is not None

        intents = dict(result.config.per_sheet_intents)
        assert "B" not in intents["Sheet2"].species_columns, (
            "Sheet2's species should match Sheet1's configuration when apply checkbox is checked"
        )


# ---------------------------------------------------------------------------
# Regression: import enablement with unified checkbox
# ---------------------------------------------------------------------------


class TestImportEnablementWithCheckbox:
    """Regression: when the unified checkbox is checked for Excel,
    import enablement must validate only the currently-viewed sheet."""

    def test_import_enabled_with_checkbox_ignores_invalid_other_sheets(self, qapp, tmp_path):
        fp = _write_xlsx(tmp_path / "multi.xlsx", {
            "Sheet1": (["time", "A", "B"], [["0", "1.0", "2.0"], ["1", "3.0", "4.0"]]),
            "Sheet2": (["time", "X", "Y"], [["0", "5.0", "6.0"], ["1", "7.0", "8.0"]]),
        })
        dlg = ImportConfigDialog(fp)

        # Switch to Sheet2, uncheck all species to make it invalid
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Sheet2":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break
        for cb in dlg._species_checkboxes:
            cb.setChecked(False)
        dlg._save_current_sheet_state()

        # Switch back to Sheet1 (valid)
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "Sheet1":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        # Without checkbox: import should be disabled (Sheet2 is invalid)
        dlg._apply_remaining_cb.setChecked(False)
        dlg._update_import_enabled()
        assert not dlg._btn_import.isEnabled(), (
            "Import must be disabled when checkbox is unchecked and Sheet2 has no species"
        )

        # With checkbox: import should be enabled (only current sheet matters)
        dlg._apply_remaining_cb.setChecked(True)
        dlg._update_import_enabled()
        assert dlg._btn_import.isEnabled(), (
            "Import must be enabled when checkbox is checked - only current sheet validated"
        )


class TestErrorHandling:
    def test_legacy_xls_is_rejected(self, qapp, tmp_path):
        path = tmp_path / "legacy.xls"
        path.write_bytes(b"legacy-xls")

        with pytest.raises(ValueError, match="Legacy \\.xls format is not supported"):
            ImportConfigDialog(str(path))

    def test_latin1_csv_shows_encoding_error_and_allows_skip_cancel(self, qapp, tmp_path):
        path = Path(tmp_path) / "latin1.csv"
        path.write_bytes(b"time,A\n0,\xff\n")

        dlg = ImportConfigDialog(str(path))

        assert "encoding error" in dlg._preview_error_label.text().lower()
        assert "utf-8" in dlg._preview_error_label.text().lower()
        assert not dlg._btn_import.isEnabled()
        assert dlg._build_result("skip").action == "skip"
        assert dlg._build_result("cancel").action == "cancel"


# ---------------------------------------------------------------------------
# 9. Sheet column mismatch validation
# ---------------------------------------------------------------------------


class TestSheetColumnMismatch:
    """Regression: config built from previewed sheet columns, not checked sheets."""

    def test_sheet_column_mismatch_preview_vs_checked_single_sheet_uses_checked_sheet_state(
        self, qapp, tmp_path, monkeypatch,
    ):
        """Check sheet A, preview sheet B: import uses sheet A's stored state."""
        fp = _write_xlsx(tmp_path / "mismatch.xlsx", {
            "SheetA": (["time", "A", "B"], [["0", "1", "2"]]),
            "SheetB": (["time", "X", "Y"], [["0", "3", "4"]]),
        })
        dlg = ImportConfigDialog(fp)

        # Uncheck SheetB, keep SheetA checked
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "SheetB":
                item.setCheckState(QtCore.Qt.CheckState.Unchecked)

        # Click SheetB to preview (species checkboxes switch to X, Y)
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "SheetB":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        dlg._on_import()

        assert dlg._result is not None
        assert dlg._result.action == "import"
        assert [plan.sheet_name for plan in dlg._result.config.plans] == ["SheetA"]
        assert dict(dlg._result.config.per_sheet_intents)["SheetA"].species_columns == ("A", "B")
        assert not criticals

    def test_sheet_column_mismatch_preview_vs_all_checked_sheets_imports_independently(
        self, qapp, tmp_path, monkeypatch,
    ):
        """Both sheets checked and incompatible: each sheet keeps its own state."""
        fp = _write_xlsx(tmp_path / "both_checked.xlsx", {
            "SheetA": (["time", "A", "B"], [["0", "1", "2"]]),
            "SheetB": (["time", "X", "Y"], [["0", "3", "4"]]),
        })
        dlg = ImportConfigDialog(fp)

        # Click SheetB to preview
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "SheetB":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        warnings: list[tuple] = []
        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "warning",
            lambda *args, **kwargs: warnings.append(args),
        )
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        dlg._on_import()

        assert dlg._result is not None
        assert [plan.sheet_name for plan in dlg._result.config.plans] == ["SheetA", "SheetB"]
        per_sheet_intents = dict(dlg._result.config.per_sheet_intents)
        assert per_sheet_intents["SheetA"].species_columns == ("A", "B")
        assert per_sheet_intents["SheetB"].species_columns == ("X", "Y")
        assert not criticals

    def test_sheet_column_mismatch_absent_when_checked_matches_preview(
        self, qapp, tmp_path, monkeypatch,
    ):
        """Check only sheet B and preview B: species X, Y are valid, import succeeds."""
        fp = _write_xlsx(tmp_path / "match.xlsx", {
            "SheetA": (["time", "A", "B"], [["0", "1", "2"]]),
            "SheetB": (["time", "X", "Y"], [["0", "3", "4"]]),
        })
        dlg = ImportConfigDialog(fp)

        # Uncheck SheetA
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "SheetA":
                item.setCheckState(QtCore.Qt.CheckState.Unchecked)

        # Click SheetB to preview
        for i in range(dlg._sheet_list.count()):
            item = dlg._sheet_list.item(i)
            if item.text() == "SheetB":
                dlg._sheet_list.setCurrentItem(item)
                dlg._on_sheet_clicked(item)
                break

        criticals: list[tuple] = []
        monkeypatch.setattr(
            QtWidgets.QMessageBox, "critical",
            lambda *args, **kwargs: criticals.append(args),
        )

        dlg._on_import()

        assert dlg._result is not None
        assert dlg._result.action == "import"
        assert set(dict(dlg._result.config.per_sheet_intents)["SheetB"].species_columns) == {"X", "Y"}
        assert list(dlg._result.config.file_intent.sheet_names) == ["SheetB"]
        assert not criticals
