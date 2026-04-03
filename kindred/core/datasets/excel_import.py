from __future__ import annotations

import datetime as dt
from typing import Iterable, Iterator, List, Mapping

from openpyxl import load_workbook

__all__ = [
    "list_sheets",
    "read_excel_sheet_rows",
]


def list_sheets(filepath: str) -> List[str]:
    """Return workbook sheet names in order."""
    workbook = _open_workbook(filepath)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    if not sheet_names:
        raise ValueError(f"Excel workbook {filepath!r} has no sheets.")
    return sheet_names


def read_excel_sheet_rows(filepath: str, sheet_name: str) -> Iterable[Mapping[str, str]]:
    """Yield sheet rows as string-valued mappings compatible with parse_csv_rows."""

    def _iter_rows() -> Iterator[Mapping[str, str]]:
        workbook = _open_workbook(filepath)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet {sheet_name!r} not found in workbook {filepath!r}.")
            yield from _read_worksheet_rows(workbook[sheet_name], filepath=filepath, sheet_name=sheet_name)
        finally:
            workbook.close()

    return _iter_rows()


def _read_worksheet_rows(
    worksheet, *, filepath: str, sheet_name: str
) -> Iterator[Mapping[str, str]]:  # noqa: ANN001
    row_iter = worksheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
        raise ValueError(f"Sheet {sheet_name!r} in workbook {filepath!r} has no rows.")
    headers = [_stringify_cell_value(value).strip() for value in header_row]
    for row in row_iter:
        values = [_stringify_cell_value(value) for value in row]
        yield dict(zip(headers, values))


def _open_workbook(filepath: str):  # noqa: ANN202
    try:
        return load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed to open Excel workbook {filepath!r}: {exc}") from exc


def _stringify_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=dt.timezone.utc)
        return repr(value.timestamp())
    if isinstance(value, dt.date):
        return repr(dt.datetime.combine(value, dt.time(), tzinfo=dt.timezone.utc).timestamp())
    if isinstance(value, dt.time):
        return repr(
            (
                value.hour * 3600
                + value.minute * 60
                + value.second
                + value.microsecond / 1_000_000
            )
        )
    if isinstance(value, float):
        return repr(value)
    return str(value)
