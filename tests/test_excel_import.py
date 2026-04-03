from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from kindred.core.datasets.csv_import import parse_csv_rows


def _save_workbook(path: Path, sheets: list[tuple[str, list[list[object]]]]) -> Path:
    workbook = Workbook()
    default_sheet = workbook.active
    first_name, first_rows = sheets[0]
    default_sheet.title = first_name
    for row in first_rows:
        default_sheet.append(list(row))

    for sheet_name, rows in sheets[1:]:
        sheet = workbook.create_sheet(title=sheet_name)
        for row in rows:
            sheet.append(list(row))

    workbook.save(path)
    workbook.close()
    return path


def test_list_sheets_returns_single_sheet_name(tmp_path: Path) -> None:
    from kindred.core.datasets.excel_import import list_sheets

    path = _save_workbook(tmp_path / "single.xlsx", [("Sheet1", [["time", "A"], [0, 1.0]])])

    assert list_sheets(str(path)) == ["Sheet1"]


def test_list_sheets_returns_named_sheets_in_order(tmp_path: Path) -> None:
    from kindred.core.datasets.excel_import import list_sheets

    path = _save_workbook(
        tmp_path / "multi.xlsx",
        [
            ("Alpha", [["time", "A"], [0, 1.0]]),
            ("Beta", [["time", "B"], [0, 2.0]]),
            ("Gamma", [["time", "C"], [0, 3.0]]),
        ],
    )

    assert list_sheets(str(path)) == ["Alpha", "Beta", "Gamma"]


@pytest.mark.parametrize("path_name", ["missing.xlsx", "corrupt.xlsx"])
def test_list_sheets_raises_for_missing_or_corrupt_workbook(tmp_path: Path, path_name: str) -> None:
    from kindred.core.datasets.excel_import import list_sheets

    path = tmp_path / path_name
    if path_name == "corrupt.xlsx":
        path.write_text("not an xlsx workbook", encoding="utf-8")

    with pytest.raises(ValueError):
        list_sheets(str(path))


def test_read_excel_sheet_rows_returns_stringified_row_dicts(tmp_path: Path) -> None:
    from kindred.core.datasets.excel_import import read_excel_sheet_rows

    path = _save_workbook(
        tmp_path / "rows.xlsx",
        [
            (
                "Data",
                [
                    ["time", "species_A", "flag"],
                    [0, 1, True],
                    [1.5, 2.25, False],
                    [2, None, None],
                ],
            )
        ],
    )

    rows = list(read_excel_sheet_rows(str(path), "Data"))

    assert rows == [
        {"time": "0", "species_A": "1", "flag": "True"},
        {"time": "1.5", "species_A": "2.25", "flag": "False"},
        {"time": "2", "species_A": "", "flag": ""},
    ]


def test_read_excel_sheet_rows_raises_for_unknown_sheet(tmp_path: Path) -> None:
    from kindred.core.datasets.excel_import import read_excel_sheet_rows

    path = _save_workbook(tmp_path / "unknown_sheet.xlsx", [("Known", [["time", "A"], [0, 1.0]])])

    with pytest.raises(ValueError, match="not found"):
        list(read_excel_sheet_rows(str(path), "Missing"))


def test_read_excel_sheet_rows_header_only_yields_no_rows(tmp_path: Path) -> None:
    from kindred.core.datasets.excel_import import read_excel_sheet_rows

    path = _save_workbook(tmp_path / "header_only.xlsx", [("Data", [["time", "A"]])])

    assert list(read_excel_sheet_rows(str(path), "Data")) == []


def test_read_excel_sheet_rows_is_lazy_and_closes_on_close(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from kindred.core.datasets import excel_import

    class _FakeSheet:
        def __init__(self) -> None:
            self.data_rows_yielded = 0

        def iter_rows(self, values_only: bool = True):
            yield ("time", "A")
            yield (0, 1.0)
            self.data_rows_yielded += 1
            yield (1, 2.0)
            self.data_rows_yielded += 1

    class _FakeWorkbook:
        def __init__(self) -> None:
            self.sheetnames = ["Data"]
            self.closed = 0
            self._sheet = _FakeSheet()

        def __getitem__(self, name: str):
            assert name == "Data"
            return self._sheet

        def close(self) -> None:
            self.closed += 1

    fake_workbook = _FakeWorkbook()
    monkeypatch.setattr(excel_import, "_open_workbook", lambda _path: fake_workbook)

    rows = excel_import.read_excel_sheet_rows(str(tmp_path / "rows.xlsx"), "Data")
    row_iter = iter(rows)

    assert fake_workbook.closed == 0
    assert fake_workbook._sheet.data_rows_yielded == 0
    assert next(row_iter) == {"time": "0", "A": "1.0"}
    assert fake_workbook.closed == 0
    assert fake_workbook._sheet.data_rows_yielded == 0
    assert next(row_iter) == {"time": "1", "A": "2.0"}
    assert fake_workbook._sheet.data_rows_yielded == 1

    rows.close()

    assert fake_workbook.closed == 1
def test_float_stringification_roundtrips_cleanly_through_parse_csv_rows(tmp_path: Path) -> None:
    from kindred.core.datasets.excel_import import read_excel_sheet_rows

    value = 1.05e-15
    path = _save_workbook(
        tmp_path / "roundtrip.xlsx",
        [("Data", [["time", "A"], [0, value], [1, 2.0]])],
    )

    rows = list(read_excel_sheet_rows(str(path), "Data"))

    assert rows[0]["A"] == repr(value)
    _time_source, payload = parse_csv_rows(rows)
    assert payload["species"]["A"][0] == value
